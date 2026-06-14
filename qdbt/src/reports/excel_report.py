"""
입찰 내 N개사 비교 Excel 보고서 생성.

compare_bid_submissions()의 결과를 받아 생성.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DB_PATH

OUTPUT_DIR = DB_PATH.parent
NAVY  = PatternFill("solid", start_color="1A3A5C")
BLUE  = PatternFill("solid", start_color="2E5C8A")
GREY  = PatternFill("solid", start_color="F3F4F6")
YELL  = PatternFill("solid", start_color="FFF9C4")
WF = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
NF = Font(name="맑은 고딕", size=10)
SF = Font(name="맑은 고딕", size=9)
BF = Font(name="맑은 고딕", size=10, bold=True)
THIN = Side(border_style="thin", color="CCCCCC")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
R    = Alignment(horizontal="right",  vertical="center")
L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _h(cell, text="", font=WF, fill=NAVY, align=C):
    cell.value = text
    cell.font  = font
    cell.fill  = fill
    cell.alignment = align
    cell.border = BOX


def _d(cell, value=None, font=NF, align=L, fmt=None):
    cell.value = value
    cell.font  = font
    cell.alignment = align
    cell.border = BOX
    if fmt:
        cell.number_format = fmt


def generate_bid_report(bid, data: dict) -> Path:
    """
    bid: sqlite3.Row  (bid_name, project_name, due_date, ...)
    data: compare_bid_submissions() 반환값
    """
    vendors = data["vendors"]
    n_v = len(vendors)

    wb = Workbook()

    # ─── 시트 1: 요약 ──────────────────────────────
    ws = wb.active
    ws.title = "요약"
    ws.merge_cells(f"A1:{get_column_letter(n_v+2)}1")
    ws["A1"] = f"입찰 비교 보고서  —  {bid['project_name']} / {bid['name']}"
    ws["A1"].font  = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill  = NAVY
    ws["A1"].alignment = C
    ws.row_dimensions[1].height = 30

    row = 3
    # 공급가액 요약
    for col, v in enumerate(vendors, start=2):
        _h(ws.cell(row, col), v, fill=BLUE)
    ws.cell(row, 1).value = "공급가액 (원)"
    ws.cell(row, 1).font = BF
    ws.cell(row, 1).alignment = L
    ws.cell(row, 1).border = BOX

    row += 1
    prices = []
    for col, v in enumerate(vendors, start=2):
        p = data["subtotals"].get(v)
        _d(ws.cell(row, col), p, align=R, fmt="#,##0")
        if p:
            prices.append((p, col))
    ws.cell(row, 1).value = "합계"
    ws.cell(row, 1).font = BF
    ws.cell(row, 1).border = BOX

    # 최저가 녹색
    if prices:
        min_col = min(prices, key=lambda x: x[0])[1]
        ws.cell(row, min_col).fill = PatternFill("solid", start_color="C6EFCE")

    # 카테고리별 소계
    row += 2
    _h(ws.cell(row, 1), "카테고리", fill=BLUE)
    for col, v in enumerate(vendors, start=2):
        _h(ws.cell(row, col), v, fill=BLUE)
    ws.row_dimensions[row].height = 22

    for cat, rows in data["categories"].items():
        row += 1
        _d(ws.cell(row, 1), cat, font=BF)
        for col, v in enumerate(vendors, start=2):
            total = data["category_totals"].get(v, {}).get(cat, 0)
            _d(ws.cell(row, col), total or None, align=R, fmt="#,##0")

    ws.column_dimensions["A"].width = 14
    for i in range(2, n_v+2):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ─── 시트 2~N: 카테고리별 항목 비교 ─────────────
    for cat, rows in data["categories"].items():
        ws2 = wb.create_sheet(cat[:20])
        ws2.merge_cells(f"A1:{get_column_letter(n_v+3)}1")
        ws2["A1"] = f"{cat}  상세 비교"
        ws2["A1"].font  = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
        ws2["A1"].fill  = BLUE
        ws2["A1"].alignment = C
        ws2.row_dimensions[1].height = 24

        # 헤더
        hdr_row = 3
        _h(ws2.cell(hdr_row, 1), "품명",   align=L)
        _h(ws2.cell(hdr_row, 2), "규격",   align=L)
        _h(ws2.cell(hdr_row, 3), "단위",   align=C)
        for col, v in enumerate(vendors, start=4):
            _h(ws2.cell(hdr_row, col), v, align=C)
        _h(ws2.cell(hdr_row, n_v+4), "최저가 업체", align=C)
        _h(ws2.cell(hdr_row, n_v+5), "최고/최저",   align=C)
        ws2.row_dimensions[hdr_row].height = 22

        for i, row_data in enumerate(rows, start=1):
            r = hdr_row + i
            _d(ws2.cell(r, 1), row_data["name"], font=NF)
            _d(ws2.cell(r, 2), row_data.get("spec") or "", font=SF, align=L)
            _d(ws2.cell(r, 3), row_data.get("unit") or "", font=SF, align=C)

            prices = []
            for col, v in enumerate(vendors, start=4):
                p = row_data["prices"].get(v)
                _d(ws2.cell(r, col), p, align=R, fmt="#,##0")
                if p:
                    prices.append((p, col))

            # 조건부 서식 (최저 녹색 / 최고 빨강)
            rng = f"{get_column_letter(4)}{r}:{get_column_letter(n_v+3)}{r}"
            gr = DifferentialStyle(fill=PatternFill("solid", start_color="C6EFCE"))
            rd = DifferentialStyle(fill=PatternFill("solid", start_color="FFC7CE"))
            ws2.conditional_formatting.add(rng,
                Rule("expression", formula=[f"AND({get_column_letter(4)}{r}=MIN(${get_column_letter(4)}${r}:${get_column_letter(n_v+3)}${r}),{get_column_letter(4)}{r}>0)"], dxf=gr))
            ws2.conditional_formatting.add(rng,
                Rule("expression", formula=[f"AND({get_column_letter(4)}{r}=MAX(${get_column_letter(4)}${r}:${get_column_letter(n_v+3)}${r}),{get_column_letter(4)}{r}>0)"], dxf=rd))

            # 최저가 업체
            if prices:
                min_p, min_col = min(prices, key=lambda x: x[0])
                # 헤더에서 업체명 찾기
                _d(ws2.cell(r, n_v+4), vendors[min_col-4], align=C, font=SF)
                max_p = max(prices, key=lambda x: x[0])[0]
                if min_p > 0:
                    ws2.cell(r, n_v+5).value = f"=MAX({get_column_letter(4)}{r}:{get_column_letter(n_v+3)}{r})/MIN({get_column_letter(4)}{r}:{get_column_letter(n_v+3)}{r})"
                    ws2.cell(r, n_v+5).number_format = '0.00"x"'
                    ws2.cell(r, n_v+5).font = SF
                    ws2.cell(r, n_v+5).alignment = C
                    ws2.cell(r, n_v+5).border = BOX

        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 6
        for col in range(4, n_v+4):
            ws2.column_dimensions[get_column_letter(col)].width = 14
        ws2.column_dimensions[get_column_letter(n_v+4)].width = 12
        ws2.column_dimensions[get_column_letter(n_v+5)].width = 8
        ws2.freeze_panes = f"D{hdr_row+1}"

    # ─── 저장 ──────────────────────────────────────
    safe_name = bid['name'].replace("/", "_").replace("\\", "_")
    out_path = OUTPUT_DIR / f"입찰비교_{safe_name}.xlsx"
    wb.save(out_path)
    return out_path
