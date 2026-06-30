# -*- coding: utf-8 -*-
"""
코드 기반 추출 엔진 (extract_by_mapping)

LLM 없이, 확인된 열 매핑으로 엑셀을 결정론적으로 추출한다.
- 열 역할(대/중/소/세분류, 품목명, 규격, 수량, 단위, 단가, 금액, 비고)을 받아
- 셀 병합 풀기 + 빈칸 상속을 코드로 처리
- 분류 열을 조립해 path 구성 (구분자 " > " 통일)
- 분류(path)와 정보(규격/단가 등)를 역할로 분리

설계: docs/QDBT_추출아키텍처_재설계.md
"""
from openpyxl import load_workbook

# 분류 역할 (순서대로 path 구성)
CAT_ROLES = ["cat1", "cat2", "cat3", "cat4", "cat5"]
# 정보 역할
INFO_ROLES = {
    "name": "name_normalized",
    "spec": "spec",
    "qty": "quantity",
    "unit": "unit",
    "price": "unit_price",
    "amount": "amount",
    "remark": "remark",
}

PATH_SEP = " > "


def _build_merge_fill(sheet):
    """병합 영역을 풀어 {(row,col): 채울값} 반환. (parse_xlsx와 동일 로직)"""
    fill = {}
    for m in sheet.merged_cells.ranges:
        top_val = sheet.cell(row=m.min_row, column=m.min_col).value
        if top_val is None:
            continue
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if r == m.min_row and c == m.min_col:
                    continue
                fill[(r, c)] = top_val
    return fill


def _to_number(v):
    """숫자 파싱 (콤마/공백/통화기호 제거). 실패 시 None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "").replace(" ", "")
    for sym in ("₩", "원", "$", "USD", "\\"):
        s = s.replace(sym, "")
    if not s:
        return None
    try:
        return float(s) if ("." in s) else int(s)
    except ValueError:
        return None


def read_grid(path, sheet_name, max_rows=None, max_cols=None):
    """엑셀 시트를 화면 표시용 격자(2차원 리스트)로 읽는다.

    인터랙티브 열 매핑 화면에서 엑셀을 그대로 보여주기 위함.
    병합 복원을 적용해 빈 병합 칸도 값이 채워진 상태로 표시.
    행/열 수는 시트 실제 크기를 따른다 (max_rows/max_cols로 상한만 선택적 지정).

    반환: {
      "grid": [[셀값, ...], ...],   # 행×열 문자열
      "n_rows": int, "n_cols": int,
      "merges": [["A1","B2"], ...],  # 병합 범위 (참고용)
    }
    """
    wb = load_workbook(path, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    merge_fill = _build_merge_fill(sheet)

    # 시트 실제 크기를 사용. max_rows/max_cols가 주어지면 상한으로만 적용.
    n_rows = sheet.max_row or 0
    n_cols = sheet.max_column or 0
    if max_rows:
        n_rows = min(n_rows, max_rows)
    if max_cols:
        n_cols = min(n_cols, max_cols)

    grid = []
    for r in range(1, n_rows + 1):
        row = []
        for c in range(1, n_cols + 1):
            v = sheet.cell(row=r, column=c).value
            if (v is None or str(v).strip() == "") and (r, c) in merge_fill:
                v = merge_fill[(r, c)]
            row.append("" if v is None else str(v))
        grid.append(row)

    merges = [[str(m.coord).split(":")[0], str(m.coord).split(":")[-1]]
              for m in sheet.merged_cells.ranges]

    wb.close()
    return {"grid": grid, "n_rows": n_rows, "n_cols": n_cols, "merges": merges}


def extract_by_mapping(path, sheet_name, column_mapping, header_row,
                       fill_down_categories=True, excluded_rows=None,
                       nego_rows=None):
    """확인된 열 매핑으로 엑셀 시트를 코드로 추출한다.

    인자:
      path: xlsx 파일 경로
      sheet_name: 시트명
      column_mapping: {열번호(1-based): 역할}
          역할: cat1~cat5(분류), name, spec, qty, unit, price, amount, remark, ignore
      header_row: 헤더 행 번호 (1-based). 이 다음 행부터 데이터.
      fill_down_categories: 분류 빈칸을 위 행에서 상속(코드 빈칸채우기). 기본 True.

    반환: {
      "items": [
        {"path": "재료비 > 기구부 > 차폐", "depth": 3,
         "name_normalized": "납 BLOCK", "spec": "순납",
         "quantity": 150, "unit": "EA", "unit_price": 64000,
         "amount": 9600000, "category": "재료비", "line_no": "R8",
         "is_category_header": False},
        ...
      ],
      "n_items": int,
      "warnings": [...],
    }
    """
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"시트 '{sheet_name}' 없음")
    sheet = wb[sheet_name]

    merge_fill = _build_merge_fill(sheet)

    def cell_val(r, c):
        """병합 복원 적용한 셀 값."""
        v = sheet.cell(row=r, column=c).value
        if (v is None or str(v).strip() == "") and (r, c) in merge_fill:
            return merge_fill[(r, c)]
        return v

    # 분류 열들 (cat1, cat2... 순)
    cat_cols = []
    for role in CAT_ROLES:
        col = next((c for c, r in column_mapping.items() if r == role), None)
        if col:
            cat_cols.append((role, col))
    # 정보 열들
    info_cols = {}
    for role, field in INFO_ROLES.items():
        col = next((c for c, r in column_mapping.items() if r == role), None)
        if col:
            info_cols[field] = col

    # 번호(seq) 열: "1", "1.1", "1.1.2" 패턴으로 계층 구성 (분류명 = 그 행 품목명)
    seq_col = next((c for c, r in column_mapping.items() if r == "seq"), None)

    items = []
    warnings = []
    excluded_rows = excluded_rows or set()
    nego_rows = nego_rows or set()
    # 분류 빈칸 상속용 (이전 행의 분류 값 기억)
    last_cat = {}  # role -> 마지막 값
    # 번호 계층용: 번호 → 그 행의 분류명 (예: "1" → "MATERIAL")
    seq_names = {}  # seq prefix -> name

    for r in range(header_row + 1, sheet.max_row + 1):
        # 사용자가 제외한 행 (subtotal 등) 스킵
        if r in excluded_rows:
            continue
        # 행 전체가 비었으면 스킵
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue

        # ── 번호(seq) 모드: 번호 패턴으로 계층 구성 ──
        if seq_col is not None:
            seq_raw = cell_val(r, seq_col)
            seq = str(seq_raw).strip() if seq_raw is not None else ""
            # 품목명/설명 (이 행의 이름)
            name_col = info_cols.get("name_normalized")
            row_name = ""
            if name_col:
                nv = cell_val(r, name_col)
                row_name = str(nv).strip() if nv is not None else ""

            # 번호가 숫자.숫자 패턴인지 (예: 1, 1.1, 1.1.2)
            import re as _re2
            if seq and _re2.match(r"^\d+(\.\d+)*$", seq):
                seq_parts = seq.split(".")
                if len(seq_parts) == 1:
                    # 최상위 번호 (예: "1") → 대분류 행. 이름을 기억하고 항목으론 스킵.
                    seq_names = {k: v for k, v in seq_names.items()}  # keep
                    seq_names[seq] = row_name or seq
                    # 대분류 행 자체는 항목 아님 (금액 없으면), 건너뜀
                    amt_col = info_cols.get("amount")
                    amt_val = _to_number(cell_val(r, amt_col)) if amt_col else None
                    if not amt_val:
                        continue  # 분류 헤더 행 → 스킵 (이름만 기억)
                    # 금액이 있으면 그대로 항목 처리 (path = 자기 이름)
                    parts = [row_name or seq]
                else:
                    # 하위 번호 (예: "1.1") → 부모들의 이름 + 자기 이름
                    parts = []
                    for d in range(1, len(seq_parts)):
                        prefix = ".".join(seq_parts[:d])
                        pname = seq_names.get(prefix, prefix)
                        parts.append(pname)
                    parts.append(row_name or seq)
            else:
                # 번호 패턴 아님 — 총계/소계 행 가능성 검사
                amt_col2 = info_cols.get("amount")
                amt2 = _to_number(cell_val(r, amt_col2)) if amt_col2 else None
                # seq 셀에 텍스트가 있으면(병합 총계행 등) 그것도 이름 후보로
                _check = ((row_name or "") + " " + (seq or "")).lower()
                _total_kw = ("total", "subtotal", "합계", "소계", "총계", "grand")
                is_total = any(k in _check for k in _total_kw)
                if amt2 and (is_total or (not seq and not row_name)):
                    continue  # 총계/소계거나 번호·이름 없이 금액만 → 스킵
                parts = [row_name] if row_name else []

            path_str = PATH_SEP.join(parts)
        else:
            # ── 기존 cat 열 모드 ──
            parts = []
            for role, col in cat_cols:
                v = cell_val(r, col)
                v = str(v).strip() if v is not None else ""
                if not v and fill_down_categories:
                    v = last_cat.get(role, "")
                if not v and parts:
                    v = parts[-1]
                if v:
                    last_cat[role] = v
                    parts.append(v)
                else:
                    if not fill_down_categories:
                        last_cat[role] = ""
            path_str = PATH_SEP.join(parts)

        # 정보 추출
        item = {
            "path": path_str,
            "depth": len(parts),
            "category": parts[0] if parts else None,
            "line_no": f"R{r}",
            "is_category_header": False,
        }
        for field, col in info_cols.items():
            v = cell_val(r, col)
            if field in ("quantity", "unit_price", "amount"):
                item[field] = _to_number(v)
            else:
                item[field] = str(v).strip() if v is not None else None

        # 차감(special nego) 행: 금액·단가를 음수로. 절댓값으로 들어와도 차감 반영.
        is_nego = (r in nego_rows)
        item["is_nego"] = is_nego
        if is_nego:
            for f in ("amount", "unit_price"):
                if item.get(f) is not None:
                    item[f] = -abs(item[f])

        # 품목명이 없으면 — 분류 헤더 행이거나 빈 행일 수 있음
        name = item.get("name_normalized")
        if not name:
            # 품목명 없고 금액만 있으면 소계 가능성 → category_header 표시
            if item.get("amount"):
                item["is_category_header"] = True
                item["name_normalized"] = parts[-1] if parts else "(소계)"
            else:
                continue  # 품목명도 금액도 없으면 스킵

        # name_raw도 채움 (없으면 normalized와 동일)
        item.setdefault("name_raw", item.get("name_normalized"))

        # 분류가 하나도 없으면(path 비어있음) 품목명을 path로.
        # 분류 없이 품목명만 있는 견적서 → 각 품목이 곧 최상위 항목.
        if not item.get("path"):
            nm = item.get("name_normalized") or item.get("name_raw") or ""
            if nm:
                item["path"] = nm
                item["depth"] = 1
                if not item.get("category"):
                    item["category"] = "기타"

        items.append(item)

    wb.close()
    return {
        "items": items,
        "n_items": len(items),
        "warnings": warnings,
    }


def suggest_column_mapping(path, sheet_name, max_scan_rows=8):
    """헤더 행을 코드로 1차 탐지해 열 매핑 후보를 제안한다.

    (LLM 헤더 인식의 코드 폴백 / 초기값. 키워드 매칭 기반.)
    반환: {"header_row": int, "mapping": {col: role}, "confidence": str}
    """
    wb = load_workbook(path, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # 역할별 헤더 키워드
    KW = {
        "seq": ["no.", "no", "번호", "순번", "항번", "item no"],
        "cat1": ["대분류"], "cat2": ["중분류"], "cat3": ["소분류"],
        "cat4": ["세분류", "세세분류"],
        "name": ["품명", "품목", "주요구성품", "주요부품", "name", "item"],
        "spec": ["규격", "사양", "spec", "remark주요"],
        "qty": ["수량", "q'ty", "qty", "수 량"],
        "unit": ["단위", "unit"],
        "price": ["단가", "unit price", "unitprice"],
        "amount": ["금액", "amount", "total", "공급가", "합계금액"],
        "remark": ["비고", "remark", "remarks"],
    }

    best_row, best_map, best_hits = None, {}, 0
    for hr in range(1, min(sheet.max_row, max_scan_rows) + 1):
        mapping = {}
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(row=hr, column=c).value
            if not v:
                continue
            vs = str(v).strip().lower()
            for role, kws in KW.items():
                if any(kw.lower() in vs for kw in kws):
                    # "unit price"는 unit이 아니라 price로 (price 우선)
                    if role == "unit" and "price" in vs:
                        continue
                    mapping[c] = role
                    break
        if len(mapping) > best_hits:
            best_hits = len(mapping)
            best_row = hr
            best_map = mapping

    wb.close()
    conf = "high" if best_hits >= 4 else ("low" if best_hits >= 2 else "none")
    return {"header_row": best_row or 1, "mapping": best_map, "confidence": conf}
