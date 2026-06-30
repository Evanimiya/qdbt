"""
XLSX 파서 — LLM 입력용 구조화 텍스트 생성

핵심 설계: LLM이 계층 구조를 추론할 수 있도록 다음 신호를 모두 보존
- 셀 좌표 (행, 열)
- 병합 셀 범위 (예: A5:F5는 병합되어 있음)
- 빈 셀 (groupby 추론용)
- 들여쓰기 (앞 공백 개수)
- 셀의 시각적 강조 (fill 색상 → 카테고리 헤더 신호)
- 셀 폰트 굵기 (헤더 신호)

출력: 시트별 텍스트 표현
- 시트 메타 (이름, 차원)
- 행 단위로 셀 값 + 위치 정보를 |로 구분
- 빈 셀은 "(empty)"로 명시
- 병합 셀 정보는 별도 섹션
"""
import openpyxl
from openpyxl import load_workbook
import json
from pathlib import Path


def cell_to_str(cell, indent_info=False):
    """셀 값을 문자열로, 들여쓰기 정보 옵션 포함"""
    v = cell.value
    if v is None or v == "":
        return "(empty)"
    if isinstance(v, str):
        # 들여쓰기 보존: 앞쪽 공백 개수 별도 표시
        leading = len(v) - len(v.lstrip(" "))
        text = v.strip()
        if indent_info and leading > 0:
            return f"[indent={leading}]{text}"
        return text
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, str) and v.startswith("="):
        # 수식 그대로
        return v
    return str(v)


def get_merged_ranges(sheet):
    """병합 셀 범위 목록"""
    return [str(rng) for rng in sheet.merged_cells.ranges]


def get_cell_format_signal(cell):
    """셀의 강조 신호: fill 색상이 있거나 bold면 헤더 가능성 높음"""
    signals = []
    try:
        if cell.font and cell.font.bold:
            signals.append("BOLD")
        fill = cell.fill
        if fill and fill.start_color and fill.start_color.rgb:
            rgb = fill.start_color.rgb
            # openpyxl은 rgb를 문자열("FFRRGGBB")로 줄 때도 있고
            # RGB 객체로 줄 때도 있어, 문자열이 아니면 문자열화 시도
            if not isinstance(rgb, str):
                rgb = str(rgb) if rgb is not None else None
            if rgb and rgb not in ("00000000", "FFFFFFFF") and len(rgb) >= 6:
                # 색상이 있음 = 강조 행 가능성
                signals.append(f"FILL={rgb[-6:]}")
    except (TypeError, AttributeError):
        # 색상/서식 정보를 못 읽어도 추출은 계속 진행 (보조 신호일 뿐)
        pass
    return ",".join(signals) if signals else ""


def build_merge_fill_map(sheet):
    """병합 영역을 풀어, 각 (행,열) 좌표에 '채워질 값'을 매핑한다.

    엑셀 병합은 좌상단 셀에만 값이 있고 나머지는 빈칸이다.
    대분류/중분류 세로 병합처럼, 병합 영역 전체에 좌상단 값을 채워두면
    각 품목 행이 자기 분류(대/중)를 알 수 있다.

    반환: {(row, col): value} — 병합으로 채워져야 할 좌표만 포함.
    """
    fill = {}
    for m in sheet.merged_cells.ranges:
        top_val = sheet.cell(row=m.min_row, column=m.min_col).value
        if top_val is None:
            continue
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if r == m.min_row and c == m.min_col:
                    continue  # 좌상단은 원래 값이 있으니 제외
                fill[(r, c)] = top_val
    return fill


def parse_sheet(sheet, max_rows=None):
    """시트를 LLM 입력용 텍스트로 변환.

    병합 셀을 미리 풀어서(좌상단 값을 병합 영역 전체에 채워서) 출력한다.
    → 대분류/중분류 세로 병합, Price 2줄 헤더 등을 LLM이 정확히 인식.
    """
    lines = []
    lines.append(f"### Sheet: {sheet.title}")
    lines.append(f"Dimensions: {sheet.dimensions}")
    lines.append(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    lines.append("")

    # 병합 셀 정보 + 복원 맵
    merged = get_merged_ranges(sheet)
    fill_map = build_merge_fill_map(sheet)
    if merged:
        lines.append("### Merged Cells (값은 아래 Cell Data에서 병합 영역 전체에 채워짐)")
        for m in merged:
            lines.append(f"  {m}")
        lines.append("")

    # 셀 데이터 (행 단위) — 병합 복원 값 적용, 간결 출력
    lines.append("### Cell Data (병합 셀은 그룹 대표값으로 채워짐)")
    lines.append("# 빈 칸은 공백, 대분류/중분류는 병합으로 여러 행에 반복됩니다.")
    n_rows = min(sheet.max_row, max_rows) if max_rows else sheet.max_row
    n_cols = sheet.max_column

    # 실제 데이터가 있는 마지막 열 찾기 (뒤쪽 빈 열 제거로 크기 축소)
    last_col = 1
    for r in range(1, n_rows + 1):
        for c in range(n_cols, last_col, -1):
            if sheet.cell(r, c).value is not None or (r, c) in fill_map:
                last_col = c
                break
    n_cols = last_col

    for r in range(1, n_rows + 1):
        row_cells = []
        signals = []
        has_content = False
        for c in range(1, n_cols + 1):
            cell = sheet.cell(r, c)
            # 병합으로 채워질 좌표면 대표값, 아니면 원래 값
            if (r, c) in fill_map:
                val = str(fill_map[(r, c)])
                has_content = True
            else:
                raw = cell.value
                if raw is None or raw == "":
                    val = ""  # 빈 칸은 짧게 (기존 "(empty)" → 빈 문자열)
                else:
                    val = cell_to_str(cell, indent_info=True)
                    has_content = True
            row_cells.append(val)
            # 포맷 시그널은 헤더 영역(상위 6행)에서만 (크기 절감)
            if r <= 6:
                sig = get_cell_format_signal(cell)
                if sig:
                    signals.append(f"{openpyxl.utils.get_column_letter(c)}{r}:{sig}")

        # 완전히 빈 행은 건너뜀 (크기 축소)
        if not has_content:
            continue

        line = f"R{r:03d} | " + " | ".join(row_cells)
        if signals:
            line += f"   <<{'  ;  '.join(signals)}>>"
        lines.append(line)

    return "\n".join(lines)


def _find_header_and_group_col(sheet, fill_map, group_keyword="중분류",
                               max_scan_rows=10):
    """헤더 행과 '묶음 기준 열'(중분류)을 탐지한다.

    반환: (header_row, group_col)  — 못 찾으면 (None, None)
    group_keyword가 들어있는 셀을 상위 행에서 찾아 그 열을 묶음 기준으로.
    """
    n_cols = sheet.max_column
    for r in range(1, min(sheet.max_row, max_scan_rows) + 1):
        for c in range(1, n_cols + 1):
            v = sheet.cell(r, c).value
            if v and group_keyword in str(v):
                return r, c
    return None, None


def parse_sheet_chunks(sheet, max_rows=30, group_keyword="중분류",
                       group_col=None):
    """시트를 묶음(중분류) 단위로 조각내어 [텍스트조각, ...] 반환.

    - 병합 복원 적용 (각 행이 대/중/소분류를 가짐)
    - 묶음 기준 열(중분류)이 바뀌는 지점에서 자름
    - 한 묶음이 max_rows 초과면 그 안에서 더 분할
    - 각 조각 = 공통 헤더 + 그 묶음의 행들

    group_col: 묶음 기준 열(1-based). None이면 group_keyword로 자동 탐지.
               자동 탐지 실패 시 max_rows 단위 기계 분할로 폴백.

    반환: {
      "chunks": [텍스트, ...],
      "group_col": 사용된 열 (None=기계분할),
      "header_row": 헤더 행,
      "n_chunks": 조각 수,
      "needs_user_input": 자동탐지 실패 여부 (True면 사용자 열 선택 필요),
    }
    """
    fill_map = build_merge_fill_map(sheet)

    # 묶음 열 결정
    header_row = None
    if group_col is None:
        header_row, group_col = _find_header_and_group_col(
            sheet, fill_map, group_keyword)

    n_rows = sheet.max_row
    n_cols = sheet.max_column

    # 셀 값 헬퍼 (병합 복원 적용)
    def cval(r, c):
        if (r, c) in fill_map:
            return str(fill_map[(r, c)])
        v = sheet.cell(r, c).value
        if v is None or v == "":
            return ""
        return cell_to_str(sheet.cell(r, c), indent_info=True)

    # 공통 헤더 텍스트 (상위 행들 = 헤더 영역)
    header_end = header_row if header_row else 4  # 헤더 못 찾으면 상위 4행 가정
    header_lines = [f"### Sheet: {sheet.title}"]
    header_lines.append("### 헤더 (모든 조각 공통)")
    for r in range(1, header_end + 1):
        cells = [cval(r, c) for c in range(1, n_cols + 1)]
        if any(cells):
            header_lines.append(f"R{r:03d} | " + " | ".join(cells))
    header_text = "\n".join(header_lines)

    # 데이터 행 범위 (헤더 다음부터)
    data_start = header_end + 1

    # ── 묶음 분할 ──
    groups = []  # [(label, [row, ...]), ...]
    if group_col:
        # 중분류 값이 바뀌는 지점으로 그룹핑
        cur_label = None
        cur_rows = []
        for r in range(data_start, n_rows + 1):
            # 빈 행 건너뜀
            row_vals = [cval(r, c) for c in range(1, n_cols + 1)]
            if not any(row_vals):
                continue
            label = cval(r, group_col)
            if label != cur_label and cur_rows:
                groups.append((cur_label, cur_rows))
                cur_rows = []
            cur_label = label
            cur_rows.append(r)
        if cur_rows:
            groups.append((cur_label, cur_rows))
    else:
        # 폴백: max_rows 단위 기계 분할
        all_rows = []
        for r in range(data_start, n_rows + 1):
            row_vals = [cval(r, c) for c in range(1, n_cols + 1)]
            if any(row_vals):
                all_rows.append(r)
        for i in range(0, len(all_rows), max_rows):
            groups.append((f"행{i+1}~", all_rows[i:i + max_rows]))

    # ── 큰 묶음 재분할 + 텍스트화 ──
    chunks = []
    for label, rows in groups:
        # max_rows 초과면 쪼갬
        for i in range(0, len(rows), max_rows):
            sub = rows[i:i + max_rows]
            lines = [header_text, f"### 데이터 (묶음: {label})"]
            for r in sub:
                cells = [cval(r, c) for c in range(1, n_cols + 1)]
                lines.append(f"R{r:03d} | " + " | ".join(cells))
            chunks.append("\n".join(lines))

    return {
        "chunks": chunks,
        "group_col": group_col,
        "header_row": header_row,
        "n_chunks": len(chunks),
        "needs_user_input": (group_col is None),
    }


def parse_xlsx_chunks(path, sheet_names=None, max_rows=30, group_keyword="중분류",
                      group_cols=None):
    """xlsx를 시트별·묶음별 조각 리스트로 반환 (청킹 추출용).

    sheet_names: 대상 시트 (None이면 전체)
    group_cols:  {시트명: 열번호} — 사용자가 지정한 묶음 열. 없으면 자동 탐지.

    반환: {
      "chunks": [텍스트조각, ...],   # 모든 시트의 조각을 이어붙임
      "needs_user_input": [시트명, ...],  # 자동탐지 실패해 사용자 선택 필요한 시트
      "sheet_info": {시트명: {n_chunks, group_col, headers}},
    }
    """
    wb_calc = load_workbook(path, data_only=True)
    targets = sheet_names if sheet_names else wb_calc.sheetnames

    all_chunks = []
    needs_input = []
    sheet_info = {}
    for name in targets:
        if name not in wb_calc.sheetnames:
            continue
        sheet = wb_calc[name]
        gcol = (group_cols or {}).get(name)
        res = parse_sheet_chunks(sheet, max_rows=max_rows,
                                 group_keyword=group_keyword, group_col=gcol)
        all_chunks.extend(res["chunks"])
        if res["needs_user_input"] and gcol is None:
            needs_input.append(name)
        sheet_info[name] = {
            "n_chunks": res["n_chunks"],
            "group_col": res["group_col"],
            "header_row": res["header_row"],
        }
    return {
        "chunks": all_chunks,
        "needs_user_input": needs_input,
        "sheet_info": sheet_info,
    }


def get_xlsx_headers(path, sheet_name, max_scan_rows=6):
    """시트의 헤더 영역(상위 행)을 사용자에게 보여주기 위해 반환.

    묶음 열 자동탐지 실패 시, 사용자가 어느 열을 묶음 기준으로 할지
    고르도록 헤더를 표로 제공.
    반환: {"columns": [(열번호, 열문자, 헤더값), ...], "preview_rows": [[...], ...]}
    """
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"columns": [], "preview_rows": []}
    sheet = wb[sheet_name]
    fill_map = build_merge_fill_map(sheet)
    n_cols = sheet.max_column

    def cval(r, c):
        if (r, c) in fill_map:
            return str(fill_map[(r, c)])
        v = sheet.cell(r, c).value
        return "" if v is None else str(v)

    # 헤더 후보: 상위 행 중 값이 가장 많은 행
    best_row, best_count = 1, 0
    for r in range(1, min(sheet.max_row, max_scan_rows) + 1):
        cnt = sum(1 for c in range(1, n_cols + 1) if cval(r, c))
        if cnt > best_count:
            best_row, best_count = r, cnt

    columns = [(c, get_column_letter(c), cval(best_row, c))
               for c in range(1, n_cols + 1) if cval(best_row, c)]
    # 미리보기: 헤더 다음 몇 행
    preview = []
    for r in range(best_row + 1, min(sheet.max_row, best_row + 4) + 1):
        preview.append([cval(r, c) for c, _, _ in columns])
    return {"columns": columns, "preview_rows": preview, "header_row": best_row}


def get_xlsx_sheet_names(path):
    """엑셀의 시트(탭) 이름 목록만 빠르게 읽는다 (전체 파싱 안 함).

    업로드 시 탭 선택 UI를 위해, 추출 전에 탭 목록만 먼저 보여줄 때 사용.
    read_only 모드로 빠르게 읽는다.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    return names


def parse_xlsx(path, max_rows_per_sheet=None, sheet_names=None):
    """XLSX 파일을 파싱.

    sheet_names: 추출할 시트 이름 목록. None이면 전체 시트(기존 동작).
                 지정 시 해당 시트만 파싱 → 불필요한 탭(일정 등) 제외로
                 입력 크기 축소, 추출 속도/안정성 향상.
    """
    wb = load_workbook(path, data_only=False)  # 수식 보존
    wb_calc = load_workbook(path, data_only=True)  # 계산값도 같이 (계산값 우선)

    # 대상 시트 결정: 지정되면 그것만(존재하는 것만), 없으면 전체
    if sheet_names:
        targets = [n for n in sheet_names if n in wb.sheetnames]
    else:
        targets = list(wb.sheetnames)

    sections = []
    sections.append(f"# File: {Path(path).name}")
    sections.append(f"# Sheets: {targets}")
    sections.append("")

    for name in targets:
        # 수식이 있는 셀은 계산값으로 표시
        sheet = wb_calc[name]
        sections.append(parse_sheet(sheet, max_rows=max_rows_per_sheet))
        sections.append("\n" + "=" * 80 + "\n")

    return "\n".join(sections)


if __name__ == "__main__":
    import sys
    files = [
        "/home/claude/sample_bid/A상사_입찰서.xlsx",
        "/home/claude/sample_bid/B제조_견적서.xlsx",
        "/home/claude/sample_bid/C공업_견적서_원본.xlsx",
    ]
    for f in files:
        out = parse_xlsx(f)
        out_path = Path("/home/claude/poc") / (Path(f).stem + ".parsed.txt")
        out_path.write_text(out, encoding="utf-8")
        print(f"  {Path(f).name} -> {out_path.name} ({len(out)} chars)")
